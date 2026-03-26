from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.infrastructure.report_stats_storage import SqliteReportStatsStorage
from app.infrastructure.report_xlsx_writer import XlsxCellLimitExceededError, write_report_xlsx


def test_write_report_xlsx_writes_sorted_rows_and_zero_filled_counts(tmp_path: Path) -> None:
    stats_path = tmp_path / "stats.sqlite"
    output_path = tmp_path / "report.xlsx"
    storage = SqliteReportStatsStorage(stats_path, reset=True)
    storage.upsert_counts(
        lemma_totals={"beta": 5, "alpha": 2},
        line_counts={
            ("beta", 2): 3,
            ("beta", 3): 2,
            ("alpha", 1): 2,
        },
    )

    write_report_xlsx(
        stats_path=stats_path,
        output_path=output_path,
        line_count=3,
        cell_char_limit=32_767,
    )

    assert output_path.is_file()
    assert not output_path.with_suffix(".xlsx.part").exists()

    workbook = load_workbook(output_path, read_only=True)
    worksheet = workbook.active
    assert list(worksheet.iter_rows(values_only=True)) == [
        ("lemma", "total_count", "counts_per_line"),
        ("alpha", 2, "2,0,0"),
        ("beta", 5, "0,3,2"),
    ]
    workbook.close()


def test_write_report_xlsx_raises_on_cell_limit_and_cleans_part_file(tmp_path: Path) -> None:
    stats_path = tmp_path / "stats.sqlite"
    output_path = tmp_path / "report.xlsx"
    storage = SqliteReportStatsStorage(stats_path, reset=True)
    storage.upsert_counts(
        lemma_totals={"alpha": 2},
        line_counts={("alpha", 1): 1, ("alpha", 2): 1},
    )

    with pytest.raises(XlsxCellLimitExceededError):
        write_report_xlsx(
            stats_path=stats_path,
            output_path=output_path,
            line_count=2,
            cell_char_limit=2,
        )

    assert not output_path.exists()
    assert not output_path.with_suffix(".xlsx.part").exists()
